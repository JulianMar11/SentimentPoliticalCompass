

import os
from openpyxl import load_workbook

import ssl
import urllib.request as ur

from bs4 import BeautifulSoup

# TODO comment in and comment in test files out
# from Data_pipe.article_content import extract_article_string
# from verify_date import range_to_dates, extract_date, check_date,

url_database_path = os.path.join("/Users/FabianFalck/Documents/[03]PotiticalCompass_PAPER/SPC/Paper/Code/Data_pipe", "url_sheet_1.xlsx")
article_database_path = os.path.join("/Users/FabianFalck/Documents/[03]PotiticalCompass_PAPER/SPC/Paper/Code/Data_pipe", "article_database.xlsx")

def retrieve_column(col, workbook_path):
    """
    Retrieves all entries of a given column col from an excel worksheet and returns them in a list

    :param col: column to retrieve
    :return: list of entries within the specified column
    """

    wb = load_workbook(filename=workbook_path)
    sheet = wb['Sheet1']
    cells = sheet[col+'2:'+col + str(sheet.max_row)]  # 2, since row 1 is column name
    entries = [str(cell[0].value) for cell in cells]  # strange tuple in tuple
    return entries


def open_URL(url):
    """
    Retrieves the webpage in a file like object
    URL must contain protocol, otherwise, the library does not work

    :param url: the url to open
    :return: tuple: 1) file like byte object (if True), empty string (if False)
                    2) True or False, depending on whether the request went through
    """
    # required, otherwise SSL CERTIFICATE error
    context = ssl._create_unverified_context()

    try:
        html = ur.urlopen(url, context=context)
        print("Opened URL: ", url)
        return html, True
    except:
        print("NOT opened URL: ", url)
        return "", False

def read_byte_content(file):
    return file.read()  # type: <class 'bytes'>

def decode_to_string(file):
    """
    makes the byte object into a proper string

    :param byte_object:
    :return: raw string of the website
    """
    # print(file)
    # trying to get the Content-Type from the server, if provided. Otherwise, take a bet.
    try:
        charset = file.info().get_content_charset()
    except:
        charset = 'utf-8'
    return file.read().decode(charset)

def soup_from_file(file):
    return BeautifulSoup(read_byte_content(file))


# content of main function is part of all_urls
if __name__ == "__main__":
    print("Do nothing")
    pass