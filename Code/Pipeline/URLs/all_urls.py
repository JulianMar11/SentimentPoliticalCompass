
import sys
sys.path.append("../")

from openpyxl import load_workbook
import uuid

from Data_pipe.text import retrieve_column
from Data_pipe.text import open_URL
from Data_pipe.all_pages import all_pages
from Data_pipe.verify_url import verify_file

url_database_path = "/Users/FabianFalck/Documents/[03]PotiticalCompass_PAPER/Data/urls_3.xlsx"
all_urls_database_path = "/Users/FabianFalck/Documents/[03]PotiticalCompass_PAPER/Data/all_urls.xlsx"

if __name__ == "__main__":

    # finding all pages of the articles found (not just the page found)
    # input
    search_terms = retrieve_column("A", url_database_path)
    newspaper_domains = retrieve_column("B", url_database_path)
    time_periods = retrieve_column("C", url_database_path)
    response_ids = retrieve_column("D", url_database_path)
    urls = retrieve_column("E", url_database_path)

    # make proper tuples
    tuples = [(search_term, newspaper_domain, time_period, response_id, url) for search_term, newspaper_domain, time_period, response_id, url in zip(search_terms, newspaper_domains, time_periods, response_ids, urls)]

    # output
    wb = load_workbook(filename=all_urls_database_path)
    sheet = wb['Sheet1']

    for t, tup in enumerate(tuples):
        search_term, newspaper_domain, time_period, response_id, url = tup

        # generate a unique identifier for this article
        article_id = str(uuid.uuid4())  # str(uuid.uuid4()) generates a unique ID

        # write the article page found also into the new workbook
        row = (search_term, newspaper_domain, time_period, response_id, url, article_id)  # response_id without the "_add" marker, since it was found during the search engine search
        sheet.append(row)

        # save the worksheet every save_every
        save_every = 5
        if t % save_every == 0:
            wb.save(all_urls_database_path)

        # get all possible pages of the article (page 1, page 2, etc.)
        pages = all_pages(url, newspaper_domain)
        for page_url in pages:
            # retrieve response
            file, opened = open_URL(page_url)
            if opened:
                # response was proper
                # Verify if the content is not suggesting that the page does not exist
                check = verify_file(file, newspaper_domain)
                if check:
                    # append that page as well
                    row = (search_term, newspaper_domain, time_period, response_id + "_add", page_url, article_id)  # "_add" marks that the page was not found as part of the initial searches
                    sheet.append(row)
                    wb.save(all_urls_database_path)
                else:
                    # page check was unsuccesfull -> content suggests that page does not exist
                    break  # breaks the innermost for loop, i.e. the pages loop
            else:
                # page was not found in request, assumption that further counting pages will also not exist
                break  # breaks the innermost for loop, i.e. the pages loop




# insert at the right position

    # INPUT
    search_terms = retrieve_column("A", url_database_path)
    newspaper_domains = retrieve_column("B", url_database_path)
    time_periods = retrieve_column("C", url_database_path)
    response_ids = retrieve_column("D", url_database_path)
    urls = retrieve_column("E", url_database_path)

    # make proper tuples
    tuples = [(search_term, newspaper_domain, time_period, response_id, url) for
              search_term, newspaper_domain, time_period, response_id, url in
              zip(search_terms, newspaper_domains, time_periods, response_ids, urls)]

    # some tests
    print(urls[1])
    file, verified = open_URL(urls[0])
    print(file)
    print(read_byte_content(file))
    print(decode_to_string(file))  # type: <class 'str'>

    # OUTPUT
    wb = load_workbook(filename=article_database_path)
    sheet = wb['Sheet1']

    for tup in tuples:
        search_term, newspaper_domain, time_period, response_id, url = tup
        article_string = ""
        # retrieve html content
        file, verified = open_URL(url)
        if verified:
            # make the soup object -> constructor call
            soup = BeautifulSoup(read_byte_content(file))
            article_string = extract_article_string(soup)

            row = (search_term, newspaper_domain, time_period, response_id, url, article_string)
            sheet.append(row)
            wb.save(article_database_path)
        else:
            article_string = "request failed"
            # no date verified in this case, since anyway useless
            # TODO leave else case away? or write to other list, since should work?





#

# verify date # TODO untested when plugged together
start, end = range_to_dates(time_period)
date_object = extract_date(newspaper_domain, soup)
# if check_date(date_object, start, end):
#     # date verified