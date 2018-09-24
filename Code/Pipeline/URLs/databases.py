
import os
from openpyxl import load_workbook

entity_database_path = os.path.join("/Users/FabianFalck/Documents/[03]PotiticalCompass_PAPER/SPC/Paper/Code/Data_pipe", "entities_without_duplicates.xlsx")
newspaper_database_path = os.path.join("/Users/FabianFalck/Documents/[03]PotiticalCompass_PAPER/SPC/Paper/Code/Data_pipe", "newspaper_database.xlsx")


def retrieve_entities(only_search_terms):
    """
    Extracts all entities in the worksheet database into a list of strings.

    :return: Entities as list of strings
    """

    wb = load_workbook(filename=entity_database_path)
    sheet = wb['Paper']

    # get all entities in the worksheet
    entity_cells = sheet['A2:A' + str(sheet.max_row)]  # 2, since row 1 is column name
    entities = [str(cell[0].value) for cell in entity_cells]  # strange tuple in tuple

    # get the column search term
    search_term_cells = sheet['D2:D' + str(sheet.max_row)]  # 2, since row 1 is column name
    search_terms = [cell[0].value for cell in search_term_cells]  # strange tuple in tuple

    # get the column importance
    importance_cells = sheet['E2:E' + str(sheet.max_row)]  # 2, since row 1 is column name
    importance_terms = [cell[0].value for cell in importance_cells]  # strange tuple in tuple

    # print(type(search_terms[0]))
    # print(search_terms[0])

    print("before search_terms: ", len(entities))
    # get only those entities that have search_term set to 1
    if only_search_terms:
        entities = [entity for i, entity in enumerate(entities) if search_terms[i] == 1]
        importances = [importance for i, importance in enumerate(importance_terms) if search_terms[i] == 1]
    elif not only_search_terms:
        entities = [entity for i, entity in enumerate(entities) if search_terms[i] == 0]
        importances = [importance for i, importance in enumerate(importance_terms) if search_terms[i] == 0]
    print("after search_terms: ", len(entities))

    # print("before unique: ", len(entities))  # TODO deprecated: duplicates removed in excel
    # get only unique entities (only process them once)
    # entities = set(entities)
    # entities = list(entities)
    # print("after unique: ", len(entities))

    print("number of entities as search terms: ", len(entities))  # should be ~ 1050 with all search terms
    print("number of importances as search terms: ", len(importances))
    assert len(entities) == len(importances)

    return entities, importances


def retrieve_newspaper_domains():
    """
    Extracts all newspaper domains in the worksheet database into a list of strings.

    :return: newspaper domains as list of strings
    """

    wb = load_workbook(filename=newspaper_database_path)
    sheet = wb['Sheet1']
    newspaper_cells = sheet['A2:A' + str(sheet.max_row)]  # 2, since row 1 is column name
    newspapers = [str(cell[0].value) for cell in newspaper_cells]  # strange tuple in tuple

    print("number of newspapers: ", len(newspapers))

    return newspapers


if __name__ == "__main__":
    print(retrieve_entities(only_search_terms=True))
    print(retrieve_newspaper_domains())